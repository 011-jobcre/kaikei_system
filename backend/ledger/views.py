# =========================================================
# Ledger Views
# =========================================================

import csv
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import models
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.views import View
from django.views.generic import TemplateView

from master.models import KanjoKamokuMaster, TorihikiSakiMaster
from journal.models import ShiwakeMeisai


# =========================================================
# General Ledger Views
# =========================================================


class SokanjoMotochoView(LoginRequiredMixin, View):
    """General Ledger (Sokanjo Motocho) View"""

    template_name = "ledger/sokanjo_motocho.html"

    def get(self, request):
        # Only level-4 active accounts are selectable
        kamoku_list = KanjoKamokuMaster.objects.filter(level=4, is_active=True).order_by("code")

        selected_kamoku = None
        meisai_list = []
        kari_total = Decimal("0")
        kashi_total = Decimal("0")
        zandaka = Decimal("0")

        kamoku_id = request.GET.get("kamoku_id", "")
        date_from_str = request.GET.get("date_from", "")
        date_to_str = request.GET.get("date_to", "")

        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None

        if kamoku_id:
            try:
                selected_kamoku = KanjoKamokuMaster.objects.get(pk=kamoku_id, level=4)
                qs = (
                    ShiwakeMeisai.objects.filter(kamoku=selected_kamoku)
                    .select_related("denpyo", "bumon", "zei_kubun")
                    .order_by("denpyo__date", "denpyo__denpyo_no")
                )
                if date_from:
                    qs = qs.filter(denpyo__date__gte=date_from)
                if date_to:
                    qs = qs.filter(denpyo__date__lte=date_to)

                # Running balance calculation
                running = Decimal("0")
                for m in qs:
                    if m.kari_kashi == "KA":
                        kari_total += m.kingaku
                        if selected_kamoku.is_kari_zandaka:
                            running += m.kingaku
                        else:
                            running -= m.kingaku
                    else:
                        kashi_total += m.kingaku
                        if selected_kamoku.is_kari_zandaka:
                            running -= m.kingaku
                        else:
                            running += m.kingaku
                    meisai_list.append(
                        {
                            "meisai": m,
                            "running_balance": running,
                        }
                    )
                zandaka = running
            except KanjoKamokuMaster.DoesNotExist:
                pass

        return render(
            request,
            self.template_name,
            {
                "kamoku_list": kamoku_list,
                "selected_kamoku": selected_kamoku,
                "selected_kamoku_id": kamoku_id,
                "meisai_list": meisai_list,
                "kari_total": kari_total,
                "kashi_total": kashi_total,
                "zandaka": zandaka,
                "date_from": date_from_str,
                "date_to": date_to_str,
                "date_from_obj": date_from,
                "date_to_obj": date_to,
            },
        )


# =========================================================
# Sub-ledger Views
# =========================================================


class HojoMotochoView(LoginRequiredMixin, View):
    """Sub-ledger (Hojo Motocho) View"""

    template_name = "ledger/hojo_motocho.html"

    def get(self, request):
        kamoku_list = KanjoKamokuMaster.objects.filter(level=4, is_active=True).order_by("code")

        torihikisaki_list = TorihikiSakiMaster.objects.filter(is_active=True).order_by("name")

        selected_kamoku = None
        selected_torihikisaki = None
        meisai_list = []
        kari_total = Decimal("0")
        kashi_total = Decimal("0")
        zandaka = Decimal("0")

        kamoku_id = request.GET.get("kamoku_id", "")
        torihikisaki_id = request.GET.get("torihikisaki_id", "")
        date_from_str = request.GET.get("date_from", "")
        date_to_str = request.GET.get("date_to", "")

        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None

        if kamoku_id and torihikisaki_id:
            try:
                selected_kamoku = KanjoKamokuMaster.objects.get(pk=kamoku_id, level=4)
                selected_torihikisaki = TorihikiSakiMaster.objects.get(pk=torihikisaki_id)
                qs = (
                    ShiwakeMeisai.objects.filter(kamoku=selected_kamoku, torihikisaki=selected_torihikisaki)
                    .select_related("denpyo", "bumon", "zei_kubun", "torihikisaki")
                    .order_by("denpyo__date", "denpyo__denpyo_no")
                )
                if date_from:
                    qs = qs.filter(denpyo__date__gte=date_from)
                if date_to:
                    qs = qs.filter(denpyo__date__lte=date_to)

                # Running balance calculation
                running = Decimal("0")
                for m in qs:
                    if m.kari_kashi == "KA":
                        kari_total += m.kingaku
                        if selected_kamoku.is_kari_zandaka:
                            running += m.kingaku
                        else:
                            running -= m.kingaku
                    else:
                        kashi_total += m.kingaku
                        if selected_kamoku.is_kari_zandaka:
                            running -= m.kingaku
                        else:
                            running += m.kingaku
                    meisai_list.append(
                        {
                            "meisai": m,
                            "running_balance": running,
                        }
                    )
                zandaka = running
            except (KanjoKamokuMaster.DoesNotExist, TorihikiSakiMaster.DoesNotExist):
                pass

        return render(
            request,
            self.template_name,
            {
                "kamoku_list": kamoku_list,
                "torihikisaki_list": torihikisaki_list,
                "selected_kamoku": selected_kamoku,
                "selected_torihikisaki": selected_torihikisaki,
                "selected_kamoku_id": kamoku_id,
                "selected_torihikisaki_id": torihikisaki_id,
                "meisai_list": meisai_list,
                "kari_total": kari_total,
                "kashi_total": kashi_total,
                "zandaka": zandaka,
                "date_from": date_from_str,
                "date_to": date_to_str,
                "date_from_obj": date_from,
                "date_to_obj": date_to,
            },
        )


# =========================================================
# Cash Book Views
# =========================================================


class GenkinSuitouchoView(LoginRequiredMixin, View):
    """Cash Book (Genkin Suitoucho) View"""

    template_name = "ledger/genkin_suitoucho.html"

    def get(self, request):
        selected_kamoku = None
        meisai_list = []
        kari_total = Decimal("0")
        kashi_total = Decimal("0")
        zandaka = Decimal("0")

        date_from_str = request.GET.get("date_from", "")
        date_to_str = request.GET.get("date_to", "")

        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None

        try:
            # Assuming '現金' is the exact name or part of the name for Cash account
            selected_kamoku = KanjoKamokuMaster.objects.filter(name__contains="現金", level=4, is_active=True).first()

            if selected_kamoku:
                qs = (
                    ShiwakeMeisai.objects.filter(kamoku=selected_kamoku)
                    .select_related("denpyo", "bumon", "zei_kubun", "torihikisaki")
                    .order_by("denpyo__date", "denpyo__denpyo_no")
                )
                if date_from:
                    qs = qs.filter(denpyo__date__gte=date_from)
                if date_to:
                    qs = qs.filter(denpyo__date__lte=date_to)

                # Running balance calculation
                running = Decimal("0")
                for m in qs:
                    if m.kari_kashi == "KA":  # 入金 (Receipt)
                        kari_total += m.kingaku
                        running += m.kingaku
                    else:  # 出金 (Payment)
                        kashi_total += m.kingaku
                        running -= m.kingaku

                    meisai_list.append(
                        {
                            "meisai": m,
                            "running_balance": running,
                        }
                    )
                zandaka = running
        except Exception:
            pass

        return render(
            request,
            self.template_name,
            {
                "selected_kamoku": selected_kamoku,
                "meisai_list": meisai_list,
                "kari_total": kari_total,
                "kashi_total": kashi_total,
                "zandaka": zandaka,
                "date_from": date_from_str,
                "date_to": date_to_str,
                "date_from_obj": date_from,
                "date_to_obj": date_to,
            },
        )


# =========================================================
# Trial Balance Views
# =========================================================


class ZandakaShisanhyouView(LoginRequiredMixin, View):
    """Trial Balance (Zandaka Shisanhyou) View"""

    template_name = "ledger/zandaka_shisanhyou.html"

    def get(self, request):
        date_to_str = request.GET.get("date_to", "")
        date_to = parse_date(date_to_str) if date_to_str else None
        kamoku_list = KanjoKamokuMaster.objects.filter(level=4, is_active=True).order_by("code")

        rows = []
        total_kari = Decimal("0")
        total_kashi = Decimal("0")

        for kamoku in kamoku_list:
            qs = ShiwakeMeisai.objects.filter(kamoku=kamoku)
            if date_to:
                qs = qs.filter(denpyo__date__lte=date_to)

            agg = qs.aggregate(
                kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
                kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
            )
            kari = agg["kari"] or Decimal("0")
            kashi = agg["kashi"] or Decimal("0")

            if kari == 0 and kashi == 0:
                continue  # Skip accounts with no activity

            if kamoku.is_kari_zandaka:
                zandaka = kari - kashi
            else:
                zandaka = kashi - kari

            rows.append(
                {
                    "kamoku": kamoku,
                    "kari_hassei": kari,
                    "kashi_hassei": kashi,
                    "zandaka": zandaka,
                }
            )
            total_kari += kari
            total_kashi += kashi

        return render(
            request,
            self.template_name,
            {
                "rows": rows,
                "total_kari": total_kari,
                "total_kashi": total_kashi,
                "date_to": date_to_str,
                "date_to_obj": date_to,
                "balanced": total_kari == total_kashi,
            },
        )


# =========================================================
# Trial Balance Export Views
# =========================================================


class ZandakaExportView(LoginRequiredMixin, View):
    """Trial Balance Export (CSV / Excel) View"""

    def get(self, request, fmt):
        date_to_str = request.GET.get("date_to", "")
        date_to = parse_date(date_to_str) if date_to_str else None
        label = date_to.strftime("%Y/%m/%d") if date_to else "All periods"

        kamoku_list = KanjoKamokuMaster.objects.filter(level=4, is_active=True).order_by("code")

        rows = []
        total_kari = Decimal("0")
        total_kashi = Decimal("0")

        for kamoku in kamoku_list:
            qs = ShiwakeMeisai.objects.filter(kamoku=kamoku)
            if date_to:
                qs = qs.filter(denpyo__date__lte=date_to)
            agg = qs.aggregate(
                kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
                kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
            )
            kari = agg["kari"] or Decimal("0")
            kashi = agg["kashi"] or Decimal("0")
            if kari == 0 and kashi == 0:
                continue
            zandaka = (kari - kashi) if kamoku.is_kari_zandaka else (kashi - kari)
            rows.append((kamoku.code, kamoku.name, kamoku.get_taisha_kubun_display(), kari, kashi, zandaka))
            total_kari += kari
            total_kashi += kashi

        if fmt == "csv":
            return self._export_csv(label, rows, total_kari, total_kashi)
        return self._export_excel(label, rows, total_kari, total_kashi)

    def _export_csv(self, label, rows, total_kari, total_kashi):
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="trial_balance.csv"'
        writer = csv.writer(response)
        writer.writerow(["Trial Balance", label])
        writer.writerow([])
        writer.writerow(["Account Code", "Account Name", "Category", "Debit Hassei", "Credit Hassei", "Balance"])
        for code, name, kubun, kari, kashi, zandaka in rows:
            writer.writerow([code, name, kubun, int(kari), int(kashi), int(zandaka)])
        writer.writerow([])
        writer.writerow(["", "Total", "", int(total_kari), int(total_kashi), ""])
        return response

    def _export_excel(self, label, rows, total_kari, total_kashi):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "残高試算表"

        ws.merge_cells("A1:F1")
        ws["A1"] = f"残高試算表 {label}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Account Code", "Account Name", "Category", "Debit Hassei", "Credit Hassei", "Balance"]
        fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True)
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=hdr)
            cell.fill = fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (code, name, kubun, kari, kashi, zandaka) in enumerate(rows, 4):
            ws.cell(row=row_idx, column=1, value=code)
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=kubun)
            for col, val in [(4, kari), (5, kashi), (6, zandaka)]:
                cell = ws.cell(row=row_idx, column=col, value=int(val))
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

        # Totals row
        t_row = len(rows) + 5
        ws.cell(row=t_row, column=2, value="Total").font = Font(bold=True)
        for col, val in [(4, total_kari), (5, total_kashi)]:
            cell = ws.cell(row=t_row, column=col, value=int(val))
            cell.number_format = "#,##0"
            cell.font = Font(bold=True, underline="double")
            cell.alignment = Alignment(horizontal="right")

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=8)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="trial_balance.xlsx"'
        return response
