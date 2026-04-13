import calendar
import csv
import io
from datetime import date as date_type
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib.auth.mixins import LoginRequiredMixin

# =========================================================
# Report Models
# =========================================================

from django.db import models
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.views import View

from journal.models import ShiwakeMeisai
from master.models import KanjoKamokuMaster, ZeiMaster


def _aggregate_by_kubun(kubun_list, date_to=None):
    """Aggregate debit/credit occurrences for the given account-type group and return balance (zandaka)."""
    qs = ShiwakeMeisai.objects.filter(kamoku__taisha_kubun__in=kubun_list)
    if date_to:
        qs = qs.filter(denpyo__date__lte=date_to)
    agg = qs.aggregate(
        kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
        kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
    )
    return agg["kari"] or Decimal("0"), agg["kashi"] or Decimal("0")


# =========================================================
# Balance Sheet（B/S）Views
# =========================================================


class BalanceSheetView(LoginRequiredMixin, View):
    """Balance Sheet（B/S）View"""

    template_name = "report/balance_sheet.html"

    def get(self, request):
        date_to_str = request.GET.get("date_to", "")
        date_to = parse_date(date_to_str) if date_to_str else None

        # Assets
        shisan_kamoku = KanjoKamokuMaster.objects.filter(taisha_kubun="SHISAN", level=4, is_active=True).order_by(
            "code"
        )
        shisan_rows = []
        total_shisan = Decimal("0")
        for k in shisan_kamoku:
            qs = ShiwakeMeisai.objects.filter(kamoku=k)
            if date_to:
                qs = qs.filter(denpyo__date__lte=date_to)
            agg = qs.aggregate(
                kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
                kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
            )
            z = (agg["kari"] or Decimal("0")) - (agg["kashi"] or Decimal("0"))
            if z != 0:
                shisan_rows.append({"kamoku": k, "zandaka": z})
                total_shisan += z

        # Liabilities
        fusai_rows, total_fusai = self._get_rows("FUSAI", date_to)
        # Equity
        junshisan_rows, total_junshisan = self._get_rows("JUNSHISAN", date_to)

        return render(
            request,
            self.template_name,
            {
                "shisan_rows": shisan_rows,
                "total_shisan": total_shisan,
                "fusai_rows": fusai_rows,
                "total_fusai": total_fusai,
                "junshisan_rows": junshisan_rows,
                "total_junshisan": total_junshisan,
                "total_fusai_junshisan": total_fusai + total_junshisan,
                "balanced": total_shisan == total_fusai + total_junshisan,
                "date_to": date_to_str,
                "date_to_obj": date_to,
            },
        )

    def _get_rows(self, taisha_kubun, date_to):
        kamoku_qs = KanjoKamokuMaster.objects.filter(taisha_kubun=taisha_kubun, level=4, is_active=True).order_by(
            "code"
        )
        rows = []
        total = Decimal("0")
        for k in kamoku_qs:
            qs = ShiwakeMeisai.objects.filter(kamoku=k)
            if date_to:
                qs = qs.filter(denpyo__date__lte=date_to)
            agg = qs.aggregate(
                kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
                kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
            )
            z = (agg["kashi"] or Decimal("0")) - (agg["kari"] or Decimal("0"))
            if z != 0:
                rows.append({"kamoku": k, "zandaka": z})
                total += z
        return rows, total


# =========================================================
# Income Statement（P/L）Views
# =========================================================


class IncomeStatementView(LoginRequiredMixin, View):
    """Income Statement（P/L）View"""

    template_name = "report/income_statement.html"

    def get(self, request):
        date_from_str = request.GET.get("date_from", "")
        date_to_str = request.GET.get("date_to", "")
        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None

        def get_rows(taisha_kubun, is_kari_zandaka):
            qs_k = KanjoKamokuMaster.objects.filter(taisha_kubun=taisha_kubun, level=4, is_active=True).order_by("code")
            rows = []
            total = Decimal("0")
            for k in qs_k:
                qs = ShiwakeMeisai.objects.filter(kamoku=k)
                if date_from:
                    qs = qs.filter(denpyo__date__gte=date_from)
                if date_to:
                    qs = qs.filter(denpyo__date__lte=date_to)
                agg = qs.aggregate(
                    kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
                    kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
                )
                ka = agg["kari"] or Decimal("0")
                sh = agg["kashi"] or Decimal("0")
                z = (ka - sh) if is_kari_zandaka else (sh - ka)
                if z != 0:
                    rows.append({"kamoku": k, "zandaka": z})
                    total += z
            return rows, total

        shueki_rows, total_shueki = get_rows("SHUEKI", False)
        hiyo_rows, total_hiyo = get_rows("HIYO", True)
        rieki = total_shueki - total_hiyo

        return render(
            request,
            self.template_name,
            {
                "shueki_rows": shueki_rows,
                "total_shueki": total_shueki,
                "hiyo_rows": hiyo_rows,
                "total_hiyo": total_hiyo,
                "rieki": rieki,
                "date_from": date_from_str,
                "date_to": date_to_str,
                "date_from_obj": date_from,
                "date_to_obj": date_to,
            },
        )


# =========================================================
# Export Helpers
# =========================================================


def _style_excel_header(ws, headers, row=1, fill_color="1E3A5F"):
    """Apply styles to Excel header row"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_column_width(ws):
    """Automatically adjust column width based on content"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# =========================================================
# Balance Sheet（B/S）Export
# =========================================================


class BalanceSheetExportView(LoginRequiredMixin, View):
    """Export Balance Sheet to Excel or CSV"""

    def get(self, request, fmt):
        date_to_str = request.GET.get("date_to", "")
        date_to = parse_date(date_to_str) if date_to_str else None
        label = date_to.strftime("%Y/%m/%d") if date_to else "最新"

        def get_section(taisha_kubun, is_kari):
            qs = KanjoKamokuMaster.objects.filter(taisha_kubun=taisha_kubun, level=4, is_active=True).order_by("code")
            rows = []
            for k in qs:
                meisai_qs = ShiwakeMeisai.objects.filter(kamoku=k)
                if date_to:
                    meisai_qs = meisai_qs.filter(denpyo__date__lte=date_to)
                agg = meisai_qs.aggregate(
                    kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
                    kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
                )
                z = (agg["kari"] or Decimal("0")) - (agg["kashi"] or Decimal("0"))
                if not is_kari:
                    z = -z
                if z != 0:
                    rows.append((k.code, k.name, z))
            return rows

        shisan_rows = get_section("SHISAN", True)
        fusai_rows = get_section("FUSAI", False)
        junshisan_rows = get_section("JUNSHISAN", False)

        if fmt == "csv":
            return self._export_csv(label, shisan_rows, fusai_rows, junshisan_rows)
        return self._export_excel(label, shisan_rows, fusai_rows, junshisan_rows)

    def _export_csv(self, label, shisan_rows, fusai_rows, junshisan_rows):
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="Balance_Sheet.csv"'
        writer = csv.writer(response)
        writer.writerow(["貸借対照表（B/S）", label])
        writer.writerow([])
        writer.writerow(["勘定科目コード", "勘定科目名", "貸借区分", "金額"])
        for code, name, z in shisan_rows:
            writer.writerow([code, name, "資産", int(z)])
        writer.writerow(["", "資産合計", "", int(sum(r[2] for r in shisan_rows))])
        writer.writerow([])
        for code, name, z in fusai_rows:
            writer.writerow([code, name, "負債", int(z)])
        writer.writerow(["", "負債合計", "", int(sum(r[2] for r in fusai_rows))])
        writer.writerow([])
        for code, name, z in junshisan_rows:
            writer.writerow([code, name, "純資産", int(z)])
        writer.writerow(["", "純資産合計", "", int(sum(r[2] for r in junshisan_rows))])
        return response

    def _export_excel(self, label, shisan_rows, fusai_rows, junshisan_rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"貸借対照表（B/S）{label}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["勘定科目コード", "勘定科目名", "貸借区分", "金額"]
        _style_excel_header(ws, headers, row=3)

        row_idx = 4
        sections = [
            ("資産", shisan_rows, "資産", "D9E7F5"),
            ("負債", fusai_rows, "負債", "FDF3E3"),
            ("純資産", junshisan_rows, "純資産", "E8F8E8"),
        ]
        for section_title, rows, kubun_label, color in sections:
            # Section header
            ws.merge_cells(f"A{row_idx}:D{row_idx}")
            cell = ws[f"A{row_idx}"]
            cell.value = section_title
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row_idx += 1
            subtotal = Decimal("0")
            for code, name, z in rows:
                ws.cell(row=row_idx, column=1, value=code)
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=kubun_label)
                amt_cell = ws.cell(row=row_idx, column=4, value=int(z))
                amt_cell.number_format = "#,##0"
                amt_cell.alignment = Alignment(horizontal="right")
                subtotal += z
                row_idx += 1
            # Subtotal row
            sub_cell = ws.cell(row=row_idx, column=4, value=int(subtotal))
            sub_cell.number_format = "#,##0"
            sub_cell.font = Font(bold=True)
            sub_cell.alignment = Alignment(horizontal="right")
            ws.merge_cells(f"B{row_idx}:C{row_idx}")
            ws.cell(row=row_idx, column=2).value = f"{section_title}合計"
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
            row_idx += 2

        # 合計 負債 + 合計 純資産
        total_fusai = sum(r[2] for r in fusai_rows)
        total_junshisan = sum(r[2] for r in junshisan_rows)
        total_fusai_junshisan = total_fusai + total_junshisan
        section_title = f"{sections[1][0]} + {sections[2][0]}"
        ws.merge_cells(f"B{row_idx}:C{row_idx}")
        ws.cell(row=row_idx, column=2).value = f"{section_title}合計"
        ws.cell(row=row_idx, column=2).font = Font(bold=True)
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        sub_cell = ws.cell(row=row_idx, column=4, value=int(total_fusai_junshisan))
        sub_cell.number_format = "#,##0"
        sub_cell.font = Font(bold=True)
        sub_cell.alignment = Alignment(horizontal="right")
        row_idx += 2

        _auto_column_width(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Balance_Sheet.xlsx"'
        return response


# =========================================================
# Income Statement（P/L）Export
# =========================================================


class IncomeStatementExportView(LoginRequiredMixin, View):
    """Export Income Statement to Excel or CSV"""

    def get(self, request, fmt):
        date_from_str = request.GET.get("date_from", "")
        date_to_str = request.GET.get("date_to", "")
        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None

        period_label = ""
        if date_from:
            period_label += date_from.strftime("%Y/%m/%d")
        if date_from and date_to:
            period_label += " 〜 "
        if date_to:
            period_label += date_to.strftime("%Y/%m/%d")
        if not date_from and not date_to:
            period_label = "全期間"

        def get_rows(taisha_kubun, is_kari_zandaka):
            qs = KanjoKamokuMaster.objects.filter(taisha_kubun=taisha_kubun, level=4, is_active=True).order_by("code")
            rows = []
            for k in qs:
                meisai_qs = ShiwakeMeisai.objects.filter(kamoku=k)
                if date_from:
                    meisai_qs = meisai_qs.filter(denpyo__date__gte=date_from)
                if date_to:
                    meisai_qs = meisai_qs.filter(denpyo__date__lte=date_to)
                agg = meisai_qs.aggregate(
                    kari=models.Sum("kingaku", filter=models.Q(kari_kashi="KA")),
                    kashi=models.Sum("kingaku", filter=models.Q(kari_kashi="SHI")),
                )
                ka = agg["kari"] or Decimal("0")
                sh = agg["kashi"] or Decimal("0")
                z = (ka - sh) if is_kari_zandaka else (sh - ka)
                if z != 0:
                    rows.append((k.code, k.name, z))
            return rows

        shueki_rows = get_rows("SHUEKI", False)
        hiyo_rows = get_rows("HIYO", True)
        total_shueki = sum(r[2] for r in shueki_rows)
        total_hiyo = sum(r[2] for r in hiyo_rows)
        rieki = total_shueki - total_hiyo

        if fmt == "csv":
            return self._export_csv(period_label, shueki_rows, hiyo_rows, total_shueki, total_hiyo, rieki)
        return self._export_excel(period_label, shueki_rows, hiyo_rows, total_shueki, total_hiyo, rieki)

    def _export_csv(self, label, shueki_rows, hiyo_rows, total_shueki, total_hiyo, rieki):
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="Income_Statement.csv"'
        writer = csv.writer(response)
        writer.writerow(["損益計算書（P/L）", label])
        writer.writerow([])
        writer.writerow(["勘定科目コード", "勘定科目名", "貸借区分", "金額"])
        for code, name, z in shueki_rows:
            writer.writerow([code, name, "収益", int(z)])
        writer.writerow(["", "収益合計 (A)", "", int(total_shueki)])
        writer.writerow([])
        for code, name, z in hiyo_rows:
            writer.writerow([code, name, "費用", int(z)])
        writer.writerow(["", "費用合計 (B)", "", int(total_hiyo)])
        writer.writerow([])
        writer.writerow(["", "当期純利益 (A-B)" if rieki >= 0 else "当期純損失 (A-B)", "", int(rieki)])
        return response

    def _export_excel(self, label, shueki_rows, hiyo_rows, total_shueki, total_hiyo, rieki):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income Statement"

        ws.merge_cells("A1:D1")
        ws["A1"] = f"損益計算書（P/L）{label}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["勘定科目コード", "勘定科目名", "貸借区分", "金額"]
        _style_excel_header(ws, headers, row=3)

        row_idx = 4
        for section_title, rows, kubun, color, subtotal in [
            ("収益の部", shueki_rows, "収益", "D9EDF7", total_shueki),
            ("費用の部", hiyo_rows, "費用", "FCF8E3", total_hiyo),
        ]:
            ws.merge_cells(f"A{row_idx}:D{row_idx}")
            cell = ws[f"A{row_idx}"]
            cell.value = section_title
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row_idx += 1
            for code, name, z in rows:
                ws.cell(row=row_idx, column=1, value=code)
                ws.cell(row=row_idx, column=2, value=name)
                ws.cell(row=row_idx, column=3, value=kubun)
                amt_cell = ws.cell(row=row_idx, column=4, value=int(z))
                amt_cell.number_format = "#,##0"
                amt_cell.alignment = Alignment(horizontal="right")
                row_idx += 1
            sub_label = "収益合計 (A)" if kubun == "収益" else "費用合計 (B)"
            ws.merge_cells(f"B{row_idx}:C{row_idx}")
            ws.cell(row=row_idx, column=2).value = sub_label
            ws.cell(row=row_idx, column=2).font = Font(bold=True)
            sub_cell = ws.cell(row=row_idx, column=4, value=int(subtotal))
            sub_cell.number_format = "#,##0"
            sub_cell.font = Font(bold=True)
            sub_cell.alignment = Alignment(horizontal="right")
            row_idx += 2

        # Net income/loss row
        net_label = "当期純利益 (A-B)" if rieki >= 0 else "当期純損失 (A-B)"
        net_color = "DFF0D8" if rieki >= 0 else "F2DEDE"
        ws.merge_cells(f"A{row_idx}:C{row_idx}")
        net_title = ws[f"A{row_idx}"]
        net_title.value = net_label
        net_title.font = Font(bold=True, size=12)
        net_title.fill = PatternFill(start_color=net_color, end_color=net_color, fill_type="solid")
        net_amt = ws.cell(row=row_idx, column=4, value=int(rieki))
        net_amt.number_format = "#,##0"
        net_amt.font = Font(bold=True, size=12)
        net_amt.alignment = Alignment(horizontal="right")
        net_amt.fill = PatternFill(start_color=net_color, end_color=net_color, fill_type="solid")

        _auto_column_width(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Income_Statement.xlsx"'
        return response


# =========================================================
# Tax Summary Report Views
# =========================================================


class ZeiSummaryView(LoginRequiredMixin, View):
    """Tax Summary Report: Aggregates taxable sales and purchases by tax rate during the period"""

    template_name = "report/zei_summary.html"

    def get(self, request):
        date_from_str = request.GET.get("date_from", "")
        date_to_str = request.GET.get("date_to", "")
        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None

        zei_list = ZeiMaster.objects.filter(is_active=True, tax_rate__gt=0).order_by("order_no")
        rows = []

        total_uriage_zeikomi = Decimal("0")
        total_shiire_zeikomi = Decimal("0")
        total_zei_uke = Decimal("0")
        total_zei_shiharai = Decimal("0")

        for zei in zei_list:
            # Taxable sales amount (incl. tax)
            qs_uriage = ShiwakeMeisai.objects.filter(
                zei_kubun=zei,
                kamoku__taisha_kubun="SHUEKI",
                kari_kashi="SHI",  # Sales are usually Credit (SHI)
            )
            # Taxable purchases/expenses amount (incl. tax)
            qs_shiire = ShiwakeMeisai.objects.filter(
                zei_kubun=zei,
                kamoku__taisha_kubun__in=["HIYO"],
                kari_kashi="KA",  # Expenses are usually Debit (KA)
            )

            if date_from:
                qs_uriage = qs_uriage.filter(denpyo__date__gte=date_from)
                qs_shiire = qs_shiire.filter(denpyo__date__gte=date_from)
            if date_to:
                qs_uriage = qs_uriage.filter(denpyo__date__lte=date_to)
                qs_shiire = qs_shiire.filter(denpyo__date__lte=date_to)

            uriage_zeikomi = qs_uriage.aggregate(total=models.Sum("kingaku"))["total"] or Decimal("0")
            shiire_zeikomi = qs_shiire.aggregate(total=models.Sum("kingaku"))["total"] or Decimal("0")

            if uriage_zeikomi == 0 and shiire_zeikomi == 0:
                continue

            # Calculate tax from amount (incl. tax -> excl. tax)
            rate = zei.tax_rate / 100  # e.g. 0.10
            divisor = 1 + rate

            uriage_zeigaku = (uriage_zeikomi * rate / divisor).quantize(Decimal("1"))
            shiire_zeigaku = (shiire_zeikomi * rate / divisor).quantize(Decimal("1"))
            uriage_honai = uriage_zeikomi - uriage_zeigaku
            shiire_honai = shiire_zeikomi - shiire_zeigaku

            rows.append(
                {
                    "zei": zei,
                    "uriage_zeikomi": uriage_zeikomi,
                    "uriage_honai": uriage_honai,
                    "uriage_zeigaku": uriage_zeigaku,
                    "shiire_zeikomi": shiire_zeikomi,
                    "shiire_honai": shiire_honai,
                    "shiire_zeigaku": shiire_zeigaku,
                }
            )

            total_uriage_zeikomi += uriage_zeikomi
            total_shiire_zeikomi += shiire_zeikomi
            total_zei_uke += uriage_zeigaku
            total_zei_shiharai += shiire_zeigaku

        nofu_zei = total_zei_uke - total_zei_shiharai  # Tax due = Received tax - Paid tax

        return render(
            request,
            self.template_name,
            {
                "rows": rows,
                "date_from": date_from_str,
                "date_to": date_to_str,
                "date_from_obj": date_from,
                "date_to_obj": date_to,
                "total_uriage_zeikomi": total_uriage_zeikomi,
                "total_shiire_zeikomi": total_shiire_zeikomi,
                "total_zei_uke": total_zei_uke,
                "total_zei_shiharai": total_zei_shiharai,
                "nofu_zei": nofu_zei,
            },
        )


class ZeiSummaryExportView(LoginRequiredMixin, View):
    """Tax Summary Report Export (CSV/Excel)"""

    def get(self, request, fmt):
        date_from_str = request.GET.get("date_from", "")
        date_to_str = request.GET.get("date_to", "")
        date_from = parse_date(date_from_str) if date_from_str else None
        date_to = parse_date(date_to_str) if date_to_str else None

        period_label = ""
        if date_from:
            period_label += date_from.strftime("%Y/%m/%d")
        if date_from and date_to:
            period_label += " 〜 "
        if date_to:
            period_label += date_to.strftime("%Y/%m/%d")
        if not date_from and not date_to:
            period_label = "全期間"

        zei_list = ZeiMaster.objects.filter(is_active=True, tax_rate__gt=0).order_by("order_no")
        rows = []
        total_zei_uke = Decimal("0")
        total_zei_shiharai = Decimal("0")

        for zei in zei_list:
            qs_uriage = ShiwakeMeisai.objects.filter(zei_kubun=zei, kamoku__taisha_kubun="SHUEKI", kari_kashi="SHI")
            qs_shiire = ShiwakeMeisai.objects.filter(zei_kubun=zei, kamoku__taisha_kubun__in=["HIYO"], kari_kashi="KA")

            if date_from:
                qs_uriage = qs_uriage.filter(denpyo__date__gte=date_from)
                qs_shiire = qs_shiire.filter(denpyo__date__gte=date_from)
            if date_to:
                qs_uriage = qs_uriage.filter(denpyo__date__lte=date_to)
                qs_shiire = qs_shiire.filter(denpyo__date__lte=date_to)

            uriage_zeikomi = qs_uriage.aggregate(total=models.Sum("kingaku"))["total"] or Decimal("0")
            shiire_zeikomi = qs_shiire.aggregate(total=models.Sum("kingaku"))["total"] or Decimal("0")

            if uriage_zeikomi == 0 and shiire_zeikomi == 0:
                continue

            rate = zei.tax_rate / 100
            divisor = 1 + rate
            uriage_zeigaku = (uriage_zeikomi * rate / divisor).quantize(Decimal("1"))
            shiire_zeigaku = (shiire_zeikomi * rate / divisor).quantize(Decimal("1"))
            uriage_honai = uriage_zeikomi - uriage_zeigaku
            shiire_honai = shiire_zeikomi - shiire_zeigaku

            rows.append(
                {
                    "name": zei.zei_name,
                    "rate": f"{zei.tax_rate}%",
                    "uri_komi": uriage_zeikomi,
                    "uri_nuki": uriage_honai,
                    "uri_zei": uriage_zeigaku,
                    "shi_komi": shiire_zeikomi,
                    "shi_nuki": shiire_honai,
                    "shi_zei": shiire_zeigaku,
                }
            )
            total_zei_uke += uriage_zeigaku
            total_zei_shiharai += shiire_zeigaku

        nofu_zei = total_zei_uke - total_zei_shiharai

        if fmt == "csv":
            return self._export_csv(period_label, rows, total_zei_uke, total_zei_shiharai, nofu_zei)
        return self._export_excel(period_label, rows, total_zei_uke, total_zei_shiharai, nofu_zei)

    def _export_csv(self, label, rows, total_uke, total_shi, nofu):
        response = HttpResponse(content_type="text/csv; charset=utf-8-sig")
        response["Content-Disposition"] = 'attachment; filename="Tax_Summary.csv"'
        writer = csv.writer(response)
        writer.writerow(["消費税集計表", label])
        writer.writerow([])
        writer.writerow(
            [
                "税区分",
                "税率",
                "課税売上（税込）",
                "課税売上（税抜）",
                "受取消費税額",
                "課税仕入（税込）",
                "課税仕入（税抜）",
                "支払消費税額",
            ]
        )
        for r in rows:
            writer.writerow(
                [
                    r["name"],
                    r["rate"],
                    int(r["uri_komi"]),
                    int(r["uri_nuki"]),
                    int(r["uri_zei"]),
                    int(r["shi_komi"]),
                    int(r["shi_nuki"]),
                    int(r["shi_zei"]),
                ]
            )
        writer.writerow([])
        writer.writerow(["受取消費税 (売上) 合計", int(total_uke)])
        writer.writerow(["支払消費税 (仕入) 合計", int(total_shi)])
        writer.writerow(["納付税額", int(nofu)])
        return response

    def _export_excel(self, label, rows, total_uke, total_shi, nofu):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consumption Tax Summary"

        ws.merge_cells("A1:H1")
        ws["A1"] = f"消費税集計表 ({label})"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "税区分",
            "税率",
            "課税売上（税込）",
            "課税売上（税抜）",
            "受取消費税額",
            "課税仕入（税込）",
            "課税仕入（税抜）",
            "支払消費税額",
        ]
        _style_excel_header(ws, headers, row=3, fill_color="1E3A5F")

        row_idx = 4
        for r in rows:
            ws.cell(row=row_idx, column=1, value=r["name"])
            ws.cell(row=row_idx, column=2, value=r["rate"])
            for col, val in enumerate(
                [r["uri_komi"], r["uri_nuki"], r["uri_zei"], r["shi_komi"], r["shi_nuki"], r["shi_zei"]], 3
            ):
                cell = ws.cell(row=row_idx, column=col, value=int(val))
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        row_idx += 1
        summary_rows = [
            ("受取消費税 (売上) 合計", total_uke, "D9EDF7"),
            ("支払消費税 (仕入) 合計", total_shi, "FCF8E3"),
            ("納付税額", nofu, "DFF0D8" if nofu >= 0 else "F2DEDE"),
        ]

        for label_text, val, color in summary_rows:
            ws.merge_cells(f"A{row_idx}:G{row_idx}")
            ws[f"A{row_idx}"] = label_text
            ws[f"A{row_idx}"].font = Font(bold=True)
            ws[f"A{row_idx}"].alignment = Alignment(horizontal="center")
            ws[f"A{row_idx}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            amt_cell = ws.cell(row=row_idx, column=8, value=int(val))
            amt_cell.font = Font(bold=True)
            amt_cell.number_format = "#,##0"
            amt_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row_idx += 1

        _auto_column_width(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Tax_Summary.xlsx"'
        return response
